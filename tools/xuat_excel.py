"""
xuat_excel.py
Xuat ket qua du bao ra tep Excel.

Nguyen tac: tep xuat ra phai TU TINH LAI duoc, khong phai anh chup cua con so.
San luong, cong suat dinh, he so tai deu la CONG THUC tro ve bang chi tiet 15 phut.
Ai sua mot o bức xa trong bang chi tiet thi ca bang tong hop doi theo. Neu ghi san
ket qua thi tep tro thanh mot buc anh chet, va sua o do se cho ra bang tong hop mau
thuan voi chinh no ma khong ai nhan ra.

Cac gia dinh (cong suat dat, buoc thoi gian) deu nam trong o rieng co nhan o trang
Thong tin, va cac cong thuc tro ve do -- khong viet thang 40 hay 0,25 vao cong thuc.
"""

from datetime import datetime
from pathlib import Path

BUOC_GIO = 0.25          # 15 phut = 0,25 gio
CAP_AC = 40.0            # MW -- cong suat dat phia xoay chieu

# ten trang; co dau cach nen moi tham chieu cheo trang deu phai dat trong nhay don
T_TONG = "Tong hop"
T_CHI_TIET = "Chi tiet 15 phut"
T_NGUON = "Theo nguon thoi tiet"
T_TIN = "Thong tin"

NAVY = "1F4E79"
XAM = "F2F6FA"
CAM = "C55A11"


def _kieu():
    from openpyxl.styles import Alignment, Font, PatternFill
    return {
        "dau": (Font(name="Arial", size=10, bold=True, color="FFFFFF"),
                PatternFill("solid", fgColor=NAVY),
                Alignment(horizontal="center", vertical="center", wrap_text=True)),
        "chu": Font(name="Arial", size=10),
        "dam": Font(name="Arial", size=10, bold=True),
        "nhan": Font(name="Arial", size=10, bold=True, color=NAVY),
        "mo": Font(name="Arial", size=9, color="595959"),
        "to": PatternFill("solid", fgColor=XAM),
    }


def _dat_dau(ws, tieu_de, dong=1):
    k = _kieu()
    f, fill, al = k["dau"]
    for i, t in enumerate(tieu_de, start=1):
        o = ws.cell(row=dong, column=i, value=t)
        o.font, o.fill, o.alignment = f, fill, al
    ws.row_dimensions[dong].height = 30
    ws.freeze_panes = ws.cell(row=dong + 1, column=1)


def xuat_excel(kq, duong_dan):
    """Ghi ket qua du bao ra tep .xlsx. `kq` la ket qua cua du_bao_tuong_lai()."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    k = _kieu()
    f_tieu_de = Font(name="Arial", size=13, bold=True, color=NAVY)
    f_nhap = Font(name="Arial", size=10, bold=True, color="0000FF")
    f_do = Font(name="Arial", size=10, color="C00000")
    f_do_dam = Font(name="Arial", size=10, bold=True, color="C00000")
    wb = Workbook()

    # ---------------------------------------------------------- trang Thong tin
    # Dat truoc vi cac trang khac tro ve day lay cong suat dat va buoc thoi gian
    ws4 = wb.active
    ws4.title = T_TIN
    mh = kq.get("mo_hinh") or {}
    ws4["A1"] = "Du bao cong suat nha may Fujiwara"
    ws4["A1"].font = f_tieu_de
    ws4["A2"] = "13,8634 do B - 109,2708 do D - 50 MWp mot chieu / 40 MW xoay chieu"
    ws4["A2"].font = k["mo"]

    # Hai o gia dinh. Moi cong thuc trong tep tro ve DAY chu khong viet so thang.
    gia_dinh = [
        ("Cong suat dat xoay chieu (MW)", CAP_AC),
        ("Buoc thoi gian (gio)", BUOC_GIO),
    ]
    ws4["A4"] = "Gia dinh dung trong cong thuc"
    ws4["A4"].font = k["nhan"]
    for i, (ten, gt) in enumerate(gia_dinh, start=5):
        ws4.cell(row=i, column=1, value=ten).font = k["chu"]
        o = ws4.cell(row=i, column=2, value=gt)
        o.font, o.fill = f_nhap, k["to"]
    O_CAP = f"'{T_TIN}'!$B$5"
    O_BUOC = f"'{T_TIN}'!$B$6"
    ws4.cell(row=7, column=1,
             value="(chu xanh = gia tri nhap tay, sua o day thi ca tep doi theo)"
             ).font = k["mo"]

    dong = 9
    khoi = [
        ("Mo hinh du bao cong suat", [
            ("Thuat toan", kq.get("ten_thuat_toan")),
            ("Tep mo hinh", mh.get("tep")),
            ("Huan luyen luc (UTC)", str(mh.get("huan_luyen_luc", ""))[:19].replace("T", " ")),
            ("Hoc tu so o ban ngay", mh.get("so_mau")),
            ("Pham vi du lieu hoc", " -> ".join(str(x)[:10] for x in (mh.get("pham_vi") or []))),
            ("Bien dau vao", ", ".join(kq.get("bien_dau_vao") or [])),
            ("Van tay du lieu", mh.get("van_tay")),
            ("Trang thai phe duyet", mh.get("trang_thai")),
            ("Du lieu da doi tu luc hoc",
             "CO - mo hinh van giu nguyen" if mh.get("du_lieu_da_doi") else "khong"),
        ]),
        ("Nguon du bao thoi tiet", (
            [(m["ten"], "dung") for m in kq.get("mo_hinh_dung") or []]
            + [(m["ten"], "loai - " + m["ly_do"]) for m in kq.get("mo_hinh_loai") or []]
            + [("Hop nhat bang", "trung vi giua cac mo hinh"),
               ("Lech tam ngay do duoc (phut)", kq.get("lech_phut"))])),
    ]
    for tieu_de, muc in khoi:
        ws4.cell(row=dong, column=1, value=tieu_de).font = k["nhan"]
        dong += 1
        for ten, gt in muc:
            ws4.cell(row=dong, column=1, value=ten).font = k["chu"]
            ws4.cell(row=dong, column=2, value=gt).font = k["chu"]
            dong += 1
        dong += 1

    ws4.cell(row=dong, column=1, value="CANH BAO QUAN TRONG").font = f_do_dam
    dong += 1
    for d in [
        "Day la du bao dua tren buc xa cua mo hinh thoi tiet (NWP), khong phai buc xa do duoc.",
        "Chi so NMAE cua phan danh gia qua khu duoc cham bang buc xa DO DUOC, nen no la tran",
        "cua khau doi buc xa sang cong suat -- KHONG phai sai so cua bang du bao nay.",
        "Sai so that chi biet duoc sau khi so voi so lieu SCADA cua chinh nhung ngay nay.",
    ]:
        ws4.cell(row=dong, column=1, value=d).font = k["chu"]
        dong += 1
    for c in kq.get("canh_bao") or []:
        dong += 1
        ws4.cell(row=dong, column=1, value="CANH BAO: " + c).font = f_do
    dong += 2
    ws4.cell(row=dong, column=1,
             value=f"Xuat luc {datetime.now():%d/%m/%Y %H:%M} - nguon buc xa: Open-Meteo (CC BY 4.0)"
             ).font = k["mo"]
    ws4.column_dimensions["A"].width = 38
    ws4.column_dimensions["B"].width = 62

    # ---------------------------------------------------------- trang Chi tiet
    ws2 = wb.create_sheet(T_CHI_TIET)
    _dat_dau(ws2, ["Ngay", "Gio", "Buc xa du bao (W/m2)", "Chi so troi quang",
                   "Cong suat du bao (MW)", "San luong o (MWh)", "Trang thai"])
    r = 2
    vung = {}          # ngay -> (dong dau, dong cuoi) de trang tong hop tro ve
    for n in kq["ngay"]:
        d0 = r
        ngay = datetime.strptime(n["ngay"], "%Y-%m-%d").date()
        for diem in n["diem"]:
            gio, p, ghi, kt = diem[0], diem[1], diem[2], diem[3]
            da_qua = len(diem) > 4 and diem[4]
            ws2.cell(row=r, column=1, value=ngay).number_format = "dd/mm/yyyy"
            ws2.cell(row=r, column=2, value=gio)
            ws2.cell(row=r, column=3, value=ghi).number_format = "#,##0"
            ws2.cell(row=r, column=4, value=kt).number_format = "0.000"
            ws2.cell(row=r, column=5, value=p).number_format = "#,##0.000"
            # San luong la CONG THUC, khong ghi san: sua cong suat thi san luong doi theo
            o = ws2.cell(row=r, column=6, value=f"=E{r}*{O_BUOC}")
            o.number_format = "#,##0.000"
            # Danh dau ro moc da troi qua: voi nhung moc do Open-Meteo tra ve gia tri
            # da hieu chinh theo quan trac, nen chung chinh xac hon du bao that. Khong
            # ghi ra thi se co nguoi lay ca ngay hom nay ra do do chinh xac cua mo hinh.
            ws2.cell(row=r, column=7, value="da troi qua" if da_qua else "")
            for c in range(1, 8):
                ws2.cell(row=r, column=c).font = k["chu"]
            r += 1
        vung[n["ngay"]] = (d0, r - 1)
    for c, w in zip("ABCDEFG", [12, 9, 20, 17, 20, 17, 13]):
        ws2.column_dimensions[c].width = w

    # ---------------------------------------------------------- trang Tong hop
    ws1 = wb.create_sheet(T_TONG, 0)
    # San luong tach lam ba cot chu khong gop mot. Ly do: mo hinh cho cong suat AM
    # ban dem (nha may tu dung), nen "tong san luong" co the hieu hai kieu -- chi phan
    # phat ra, hay tru ca phan tu dung. Hai cach lech nhau khoang 4 MWh mot ngay. Gop
    # mot cot thi hai nguoi doc cung mot tep se ra hai con so khac nhau ma khong ai
    # biet vi sao; tach ra thi quan he giua chung hien ngay tren mat bang.
    _dat_dau(ws1, ["Ngay", "Thu", "San luong phat (MWh)", "Tu dung ban dem (MWh)",
                   "San luong rong (MWh)", "Cong suat dinh (MW)", "Gio dat dinh",
                   "He so tai", "Chi so troi quang trung binh",
                   "Tan giua cac nguon (MWh)"])
    q = f"'{T_CHI_TIET}'"
    for i, n in enumerate(kq["ngay"], start=2):
        a, b = vung[n["ngay"]]
        ws1.cell(row=i, column=1,
                 value=datetime.strptime(n["ngay"], "%Y-%m-%d").date()
                 ).number_format = "dd/mm/yyyy"
        ws1.cell(row=i, column=2, value=n["thu"])
        ws1.cell(row=i, column=3,
                 value=f'=SUMIF({q}!F{a}:F{b},">0")').number_format = "#,##0.0"
        ws1.cell(row=i, column=4,
                 value=f'=SUMIF({q}!F{a}:F{b},"<0")').number_format = "#,##0.0"
        ws1.cell(row=i, column=5, value=f"=C{i}+D{i}").number_format = "#,##0.0"
        ws1.cell(row=i, column=6, value=f"=MAX({q}!E{a}:E{b})").number_format = "#,##0.00"
        # Gio dat dinh doc nguoc tu bang chi tiet bang INDEX/MATCH, khong ghi san --
        # neu ai sua mot o cong suat thi gio dinh phai di theo, neu khong bang nay
        # se tu mau thuan voi chinh no.
        ws1.cell(row=i, column=7,
                 value=f"=INDEX({q}!$B${a}:$B${b},MATCH(F{i},{q}!$E${a}:$E${b},0))")
        ws1.cell(row=i, column=8, value=f"=C{i}/({O_CAP}*24)").number_format = "0.0%"
        ws1.cell(row=i, column=9,
                 value=f'=IFERROR(AVERAGEIF({q}!$D${a}:$D${b},">0.001"),0)'
                 ).number_format = "0.000"
        ws1.cell(row=i, column=10, value=n.get("tan_mwh")).number_format = "#,##0.0"
        for c in range(1, 11):
            ws1.cell(row=i, column=c).font = k["chu"]
    n_cuoi = len(kq["ngay"]) + 1
    ws1.cell(row=n_cuoi + 1, column=2, value="Tong").font = k["dam"]
    for c in "CDE":
        o = ws1.cell(row=n_cuoi + 1, column=ord(c) - 64, value=f"=SUM({c}2:{c}{n_cuoi})")
        o.number_format, o.font = "#,##0.0", k["dam"]
    for j, t in enumerate([
        f"Moi con so trong bang nay la cong thuc tro ve trang '{T_CHI_TIET}'; he so tai "
        f"chia cho cong suat dat o trang '{T_TIN}' o B5. Sua bang chi tiet thi bang nay "
        "tu doi theo.",
        "San luong phat = tong cac o co cong suat duong. Tu dung ban dem = tong cac o am, "
        "la muc cong suat nen ban dem uoc luong tu chinh so lieu nha may. He so tai tinh "
        "theo san luong PHAT de khop voi con so hien tren giao dien web.",
        "Tan giua cac nguon la khoang cach giua mo hinh thoi tiet cho san luong cao nhat "
        "va thap nhat. Tan rong = ngay kho du bao, khong phai ngay it nang.",
    ], start=3):
        ws1.cell(row=n_cuoi + j, column=1, value=t).font = k["mo"]
    for c, w in zip("ABCDEFGHIJ", [12, 11, 19, 21, 19, 20, 13, 11, 27, 24]):
        ws1.column_dimensions[c].width = w

    # ---------------------------------------------------------- trang Theo nguon
    ten_mh = [m["ma"] for m in kq.get("mo_hinh_dung") or []]
    if ten_mh:
        ws3 = wb.create_sheet(T_NGUON)
        nhan = {m["ma"]: m["ten"] for m in kq["mo_hinh_dung"]}
        _dat_dau(ws3, ["Ngay", "Thu"] + [nhan[m] for m in ten_mh] + ["Trung vi (dung)"])
        for i, n in enumerate(kq["ngay"], start=2):
            ws3.cell(row=i, column=1,
                     value=datetime.strptime(n["ngay"], "%Y-%m-%d").date()
                     ).number_format = "dd/mm/yyyy"
            ws3.cell(row=i, column=2, value=n["thu"])
            for j, m in enumerate(ten_mh, start=3):
                ws3.cell(row=i, column=j,
                         value=n["e_theo_mo_hinh"].get(m)).number_format = "#,##0.0"
            c = get_column_letter(2 + len(ten_mh) + 1)
            ws3.cell(row=i, column=2 + len(ten_mh) + 1,
                     value=f"='{T_TONG}'!C{i}").number_format = "#,##0.0"
            for j in range(1, 4 + len(ten_mh)):
                ws3.cell(row=i, column=j).font = k["chu"]
        ws3.cell(row=len(kq["ngay"]) + 3, column=1,
                 value="San luong ngay neu chi dung rieng tung nguon thoi tiet. Con so "
                       "dung trong bao cao la cot Trung vi -- trung vi khong bi mot nguon "
                       "chay sai keo di, khac voi trung binh cong."
                 ).font = k["mo"]
        ws3.column_dimensions["A"].width = 12
        ws3.column_dimensions["B"].width = 11
        for j in range(3, 4 + len(ten_mh)):
            ws3.column_dimensions[get_column_letter(j)].width = 22

    duong_dan = Path(duong_dan)
    duong_dan.parent.mkdir(parents=True, exist_ok=True)
    wb.save(duong_dan)
    return duong_dan
